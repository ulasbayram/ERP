from __future__ import annotations

import cgi
import hashlib
import hmac
import json
import os
import shutil
import sqlite3
import sys
import secrets
import time
from datetime import date, datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from http import cookies
from urllib.parse import parse_qs, urlparse

import openpyxl


ROOT = Path(__file__).resolve().parent
PUBLIC = ROOT / "public"
DATA = ROOT / "data"
UPLOADS = DATA / "uploads"
ATTACHMENTS = DATA / "attachments"
DB_PATH = DATA / "erp.sqlite3"
SCHEMA = ROOT / "schema.sql"
SESSION_COOKIE = "akim_erp_session"
PASSWORD_ITERATIONS = 600_000
SESSION_DAYS = int(os.environ.get("ERP_SESSION_DAYS", "7"))
COOKIE_SECURE_MODE = os.environ.get("ERP_COOKIE_SECURE", "auto").strip().lower()
INVITE_CODE = os.environ.get("ERP_INVITE_CODE", "")
ALLOW_OPEN_REGISTRATION = os.environ.get("ERP_ALLOW_OPEN_REGISTRATION", "false").lower() == "true"
HOST = os.environ.get("ERP_HOST", "127.0.0.1")
MAX_UPLOAD_BYTES = int(os.environ.get("ERP_MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
RATE_LIMITS: dict[str, list[float]] = {}
DEFAULT_COMPANY_NAME = "Ulaş Bayram ERP"
ADMIN_ROLE = "admin"
OWNER_ROLE = "owner"
ACCOUNTANT_ROLE = "accountant"
LOG_VIEW_ROLES = {ADMIN_ROLE, OWNER_ROLE}
ASSIGNABLE_ROLES = {ADMIN_ROLE, OWNER_ROLE, ACCOUNTANT_ROLE}


class CompanyAccessError(PermissionError):
    pass


def normalize_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_key(value) -> str:
    return normalize_text(value).casefold()


def normalize_status(value) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    text = raw.casefold()
    if "ödendi" in text or "odendi" in text:
        return "paid"
    if "kısmi" in text or "kismi" in text:
        return "partial"
    if "bek" in text:
        return "pending"
    if "gec" in text:
        return "overdue"
    return raw.lower()


def turkish_title(value: str) -> str:
    lower_map = str.maketrans("IİĞÜŞÖÇ", "ıiğüşöç")
    upper_map = str.maketrans("ıiğüşöç", "IİĞÜŞÖÇ")
    text = normalize_text(value).translate(lower_map).lower()
    return " ".join(word[:1].translate(upper_map).upper() + word[1:] for word in text.split())


def clean_job_title(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "-" in text:
        text = text.split("-", 1)[1]
    text = text.strip(" -.")
    return turkish_title(text)


def to_iso(value) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    return text or None


def period_from_date(value: str | None = None) -> str:
    text = normalize_text(value)
    if text:
        return text[:7]
    return date.today().strftime("%Y-%m")


def month_bounds(period: str | None = None) -> tuple[date, date]:
    clean = normalize_text(period) or date.today().strftime("%Y-%m")
    try:
        year, month = [int(part) for part in clean[:7].split("-")]
        start = date(year, month, 1)
    except Exception:
        today = date.today()
        start = date(today.year, today.month, 1)
    next_month = date(start.year + (1 if start.month == 12 else 0), 1 if start.month == 12 else start.month + 1, 1)
    return start, next_month - timedelta(days=1)


def parse_date(value) -> date | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError:
        return None


def payroll_base_salary(monthly_salary: float, hire_date: str | None, leave_date: str | None, period: str | None = None) -> tuple[float, int]:
    start, end = month_bounds(period)
    hired = parse_date(hire_date) or start
    left = parse_date(leave_date) or end
    active_start = max(hired, start)
    active_end = min(left, end)
    if active_end < active_start:
        return 0.0, 0
    active_days = (active_end - active_start).days + 1
    calendar_days = (end - start).days + 1
    return float(monthly_salary or 0) * active_days / calendar_days, active_days


def to_float(value) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if value is None:
        return 0.0
    text = normalize_text(value).replace("₺", "").replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def read_json_body(handler) -> dict:
    length = int(handler.headers.get("content-length", "0") or 0)
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


def hash_password(password: str, salt: bytes | None = None, iterations: int = PASSWORD_ITERATIONS) -> tuple[str, str, int]:
    salt = salt or secrets.token_bytes(32)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return digest.hex(), salt.hex(), iterations


def verify_password(password: str, stored_hash: str, stored_salt: str, iterations: int) -> bool:
    digest, _, _ = hash_password(password, bytes.fromhex(stored_salt), iterations)
    return hmac.compare_digest(digest, stored_hash)


def hash_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def rate_limited(key: str, limit: int, window_seconds: int) -> bool:
    now = time.time()
    bucket = RATE_LIMITS.setdefault(key, [])
    RATE_LIMITS[key] = [stamp for stamp in bucket if now - stamp < window_seconds]
    if len(RATE_LIMITS[key]) >= limit:
        return True
    RATE_LIMITS[key].append(now)
    return False


def user_count(conn: sqlite3.Connection) -> int:
    return int(conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"])


def ensure_default_company(conn: sqlite3.Connection) -> int:
    conn.execute("INSERT OR IGNORE INTO companies(name) VALUES (?)", (DEFAULT_COMPANY_NAME,))
    row = conn.execute("SELECT id FROM companies WHERE name = ?", (DEFAULT_COMPANY_NAME,)).fetchone()
    return int(row["id"])


def list_companies(conn: sqlite3.Connection) -> list[dict]:
    return [dict(row) for row in conn.execute("SELECT id, name, tax_number, status, created_at FROM companies ORDER BY status, name")]


def list_users(conn: sqlite3.Connection) -> list[dict]:
    return [
        dict(row)
        for row in conn.execute(
            """
            SELECT u.id, u.company_id, u.email, u.full_name, u.role, u.is_active,
                   u.created_at, u.last_login_at, c.name AS company_name
            FROM users u
            LEFT JOIN companies c ON c.id = u.company_id
            ORDER BY u.created_at DESC, u.id DESC
            """
        )
    ]


def can_view_logs(user: dict | None) -> bool:
    return bool(user and user.get("role") in LOG_VIEW_ROLES)


def parse_company_id(value) -> int | None:
    try:
        company_id = int(value)
    except (TypeError, ValueError):
        return None
    return company_id if company_id > 0 else None


def resolve_company_id(conn: sqlite3.Connection, user: dict, requested=None) -> int:
    default_company_id = ensure_default_company(conn)
    if user.get("role") == ADMIN_ROLE:
        requested_id = parse_company_id(requested)
        if requested_id and conn.execute("SELECT id FROM companies WHERE id = ?", (requested_id,)).fetchone():
            return requested_id
        user_company_id = parse_company_id(user.get("company_id"))
        return user_company_id or default_company_id
    user_company_id = parse_company_id(user.get("company_id"))
    if not user_company_id:
        raise CompanyAccessError("Bu hesap henüz bir firmaya bağlanmamış. Admin ataması gerekli.")
    company = conn.execute("SELECT id, status FROM companies WHERE id = ?", (user_company_id,)).fetchone()
    if not company:
        raise CompanyAccessError("Bu hesabın bağlı olduğu firma bulunamadı.")
    if normalize_text(company["status"]) == "archived":
        raise CompanyAccessError("Bu hesabın bağlı olduğu firma arşivlenmiş. Admin kontrolü gerekli.")
    return user_company_id


def company_id_from_request(handler, payload: dict | None = None, form=None) -> int | None:
    if payload:
        found = parse_company_id(payload.get("companyId"))
        if found:
            return found
    if form is not None:
        found = parse_company_id(form.getfirst("companyId"))
        if found:
            return found
    header_value = handler.headers.get("X-Company-Id")
    return parse_company_id(header_value)


def audit_event(
    conn: sqlite3.Connection,
    company_id: int | None,
    actor: str,
    action: str,
    entity_name: str,
    entity_id: str | int | None = None,
    old_value=None,
    new_value=None,
) -> None:
    conn.execute(
        """
        INSERT INTO audit_events(company_id, actor, action, entity_name, entity_id, old_value, new_value)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_id,
            actor,
            action,
            entity_name,
            str(entity_id) if entity_id is not None else None,
            json.dumps(old_value, ensure_ascii=False) if isinstance(old_value, (dict, list)) else old_value,
            json.dumps(new_value, ensure_ascii=False) if isinstance(new_value, (dict, list)) else new_value,
        ),
    )


def registration_allowed(conn: sqlite3.Connection) -> bool:
    if user_count(conn) == 0:
        return True
    return ALLOW_OPEN_REGISTRATION or bool(INVITE_CODE)


def create_session(conn: sqlite3.Connection, user_id: int, user_agent: str, ip_address: str) -> str:
    token = secrets.token_urlsafe(48)
    expires = datetime.utcnow() + timedelta(days=SESSION_DAYS)
    conn.execute(
        """
        INSERT INTO auth_sessions(user_id, token_hash, expires_at, user_agent, ip_address)
        VALUES (?, ?, ?, ?, ?)
        """,
        (user_id, hash_token(token), expires.isoformat(timespec="seconds"), user_agent[:400], ip_address),
    )
    return token


def current_user_from_cookie(cookie_header: str | None) -> dict | None:
    if not cookie_header:
        return None
    jar = cookies.SimpleCookie()
    try:
        jar.load(cookie_header)
    except cookies.CookieError:
        return None
    morsel = jar.get(SESSION_COOKIE)
    if not morsel:
        return None
    token_digest = hash_token(morsel.value)
    now = datetime.utcnow().isoformat(timespec="seconds")
    with connect() as conn:
        conn.execute("DELETE FROM auth_sessions WHERE expires_at < ?", (now,))
        row = conn.execute(
            """
            SELECT u.id, u.company_id, u.email, u.full_name, u.role
            FROM auth_sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.token_hash = ? AND s.expires_at >= ? AND u.is_active = 1
            """,
            (token_digest, now),
        ).fetchone()
        return dict(row) if row else None


def mask_identifier(value: str, visible: int = 4) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if len(text) <= visible:
        return "*" * len(text)
    return "*" * (len(text) - visible) + text[-visible:]


def connect() -> sqlite3.Connection:
    DATA.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.OperationalError:
        pass
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def backup_database(reason: str) -> str | None:
    if not DB_PATH.exists():
        return None
    backup_dir = DATA / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{stamp}_{reason}.sqlite3"
    shutil.copy2(DB_PATH, backup_path)
    return str(backup_path)


def same_origin_allowed(headers) -> bool:
    origin = headers.get("Origin")
    if not origin:
        return True
    host = headers.get("Host", "")
    return origin in {f"http://{host}", f"https://{host}"}


def should_send_secure_cookie(headers) -> bool:
    if COOKIE_SECURE_MODE in {"1", "true", "yes", "on"}:
        return True
    if COOKIE_SECURE_MODE in {"0", "false", "no", "off"}:
        return False

    forwarded_proto = headers.get("X-Forwarded-Proto", "").split(",", 1)[0].strip().lower()
    forwarded = headers.get("Forwarded", "").lower()
    cf_visitor = headers.get("Cf-Visitor", "").lower()
    return (
        forwarded_proto == "https"
        or "proto=https" in forwarded
        or '"scheme":"https"' in cf_visitor
        or "'scheme':'https'" in cf_visitor
    )


def init_db() -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    UPLOADS.mkdir(parents=True, exist_ok=True)
    ATTACHMENTS.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript(SCHEMA.read_text(encoding="utf-8"))
        ensure_schema_migrations(conn)
        conn.execute(
            "INSERT OR IGNORE INTO companies(name) VALUES (?)",
            ("Ulaş Bayram ERP",),
        )


def ensure_schema_migrations(conn: sqlite3.Connection) -> None:
    default_company_id = ensure_default_company(conn)
    company_columns = {row["name"] for row in conn.execute("PRAGMA table_info(companies)")}
    if "status" not in company_columns:
        conn.execute("ALTER TABLE companies ADD COLUMN status TEXT NOT NULL DEFAULT 'active'")
    conn.execute("UPDATE companies SET status = 'active' WHERE status IS NULL OR status = ''")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS account_movements (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          account_kind TEXT NOT NULL,
          account_id INTEGER NOT NULL,
          movement_date TEXT NOT NULL,
          movement_type TEXT NOT NULL,
          direction TEXT NOT NULL,
          amount REAL NOT NULL DEFAULT 0,
          document_no TEXT,
          description TEXT,
          source_table TEXT,
          source_id INTEGER,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS entity_attachments (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          entity_type TEXT NOT NULL,
          entity_id INTEGER NOT NULL,
          file_name TEXT NOT NULL,
          stored_name TEXT NOT NULL,
          mime_type TEXT,
          file_size INTEGER NOT NULL DEFAULT 0,
          uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS bank_transfer_vouchers (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          transfer_date TEXT NOT NULL,
          from_account_code TEXT NOT NULL,
          to_account_code TEXT NOT NULL,
          amount REAL NOT NULL DEFAULT 0,
          description TEXT,
          source_bank_line_id INTEGER,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          FOREIGN KEY (source_bank_line_id) REFERENCES bank_statement_lines(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS employee_advances (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          employee_id INTEGER NOT NULL,
          advance_date TEXT NOT NULL,
          period TEXT NOT NULL,
          amount REAL NOT NULL DEFAULT 0,
          note TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          FOREIGN KEY (employee_id) REFERENCES employees(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS employee_overtime_entries (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          employee_id INTEGER NOT NULL,
          overtime_date TEXT NOT NULL,
          period TEXT NOT NULL,
          hours REAL NOT NULL DEFAULT 0,
          hourly_rate REAL NOT NULL DEFAULT 0,
          amount REAL NOT NULL DEFAULT 0,
          note TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          FOREIGN KEY (employee_id) REFERENCES employees(id)
        )
        """
    )
    employee_columns = {row["name"] for row in conn.execute("PRAGMA table_info(employees)")}
    employee_migrations = {
        "leave_date": "ALTER TABLE employees ADD COLUMN leave_date TEXT",
        "iban_masked": "ALTER TABLE employees ADD COLUMN iban_masked TEXT",
        "phone_masked": "ALTER TABLE employees ADD COLUMN phone_masked TEXT",
    }
    for column, statement in employee_migrations.items():
        if column not in employee_columns:
            conn.execute(statement)
    bank_columns = {row["name"] for row in conn.execute("PRAGMA table_info(bank_statement_lines)")}
    bank_migrations = {
        "match_status": "ALTER TABLE bank_statement_lines ADD COLUMN match_status TEXT NOT NULL DEFAULT 'unmatched'",
        "match_type": "ALTER TABLE bank_statement_lines ADD COLUMN match_type TEXT",
        "matched_partner_id": "ALTER TABLE bank_statement_lines ADD COLUMN matched_partner_id INTEGER",
        "matched_invoice_id": "ALTER TABLE bank_statement_lines ADD COLUMN matched_invoice_id INTEGER",
        "account_code": "ALTER TABLE bank_statement_lines ADD COLUMN account_code TEXT",
        "match_note": "ALTER TABLE bank_statement_lines ADD COLUMN match_note TEXT",
        "matched_at": "ALTER TABLE bank_statement_lines ADD COLUMN matched_at TEXT",
    }
    for column, statement in bank_migrations.items():
        if column not in bank_columns:
            conn.execute(statement)
    company_scoped_tables = [
        "users",
        "project_sites",
        "business_partners",
        "employees",
        "employee_advances",
        "employee_overtime_entries",
        "bank_statement_lines",
        "purchase_invoices",
        "payment_instruments",
        "account_movements",
        "entity_attachments",
        "bank_transfer_vouchers",
        "reference_values",
        "import_batches",
        "audit_events",
    ]
    for table in company_scoped_tables:
        columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}
        added_company_column = False
        if "company_id" not in columns:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN company_id INTEGER")
            added_company_column = True
        if table != "users" or added_company_column:
            conn.execute(f"UPDATE {table} SET company_id = ? WHERE company_id IS NULL", (default_company_id,))
    conn.execute("CREATE INDEX IF NOT EXISTS idx_project_sites_company_name ON project_sites(company_id, name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_business_partners_company_name ON business_partners(company_id, normalized_name)")
    conn.execute(
        """
        INSERT INTO account_movements(
          company_id, account_kind, account_id, movement_date, movement_type, direction, amount,
          document_no, description, source_table, source_id
        )
        SELECT p.company_id, 'partner', p.partner_id, COALESCE(p.invoice_date, date('now')), 'invoice', 'credit',
               p.gross_total, p.invoice_no, COALESCE(p.description, 'Alış faturası'),
               'purchase_invoices', p.id
        FROM purchase_invoices p
        WHERE p.partner_id IS NOT NULL
          AND p.gross_total > 0
          AND NOT EXISTS (
            SELECT 1
            FROM account_movements m
            WHERE m.account_kind = 'partner'
              AND m.account_id = p.partner_id
              AND m.company_id = p.company_id
              AND m.source_table = 'purchase_invoices'
              AND m.source_id = p.id
              AND m.movement_type = 'invoice'
          )
        """
    )
    conn.execute(
        """
        INSERT INTO account_movements(
          company_id, account_kind, account_id, movement_date, movement_type, direction, amount,
          document_no, description, source_table, source_id
        )
        SELECT p.company_id, 'partner', p.partner_id, COALESCE(p.invoice_date, date('now')), 'payment', 'debit',
               p.paid_amount, p.invoice_no, 'Fatura ödemesi',
               'purchase_invoices', p.id
        FROM purchase_invoices p
        WHERE p.partner_id IS NOT NULL
          AND p.paid_amount > 0
          AND NOT EXISTS (
            SELECT 1
            FROM account_movements m
            WHERE m.account_kind = 'partner'
              AND m.account_id = p.partner_id
              AND m.company_id = p.company_id
              AND m.source_table = 'purchase_invoices'
              AND m.source_id = p.id
              AND m.movement_type = 'payment'
          )
        """
    )


def header_map(ws) -> dict[str, int]:
    return {
        normalize_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if normalize_text(ws.cell(1, col).value)
    }


def value(ws, row: int, headers: dict[str, int], name: str):
    col = headers.get(name)
    if not col:
        return None
    return ws.cell(row, col).value


def get_or_create_partner(conn: sqlite3.Connection, name: str, partner_type: str = "vendor", company_id: int | None = None) -> int | None:
    clean = normalize_text(name)
    if not clean:
        return None
    normalized = normalize_key(clean)
    company_id = company_id or ensure_default_company(conn)
    row = conn.execute(
        "SELECT id FROM business_partners WHERE company_id = ? AND normalized_name = ?",
        (company_id, normalized),
    ).fetchone()
    if row:
        return int(row["id"])
    cur = conn.execute(
        """
        INSERT INTO business_partners(company_id, name, partner_type, normalized_name)
        VALUES (?, ?, ?, ?)
        """,
        (company_id, clean, partner_type, normalized),
    )
    return int(cur.lastrowid)


def get_or_create_site(conn: sqlite3.Connection, name: str, company_id: int | None = None) -> int | None:
    clean = normalize_text(name)
    if not clean:
        return None
    company_id = company_id or ensure_default_company(conn)
    row = conn.execute("SELECT id FROM project_sites WHERE company_id = ? AND name = ?", (company_id, clean)).fetchone()
    if row:
        return int(row["id"])
    cur = conn.execute("INSERT INTO project_sites(company_id, name) VALUES (?, ?)", (company_id, clean))
    return int(cur.lastrowid)


def add_account_movement(
    conn: sqlite3.Connection,
    company_id: int,
    account_kind: str,
    account_id: int,
    movement_date: str | None,
    movement_type: str,
    direction: str,
    amount: float,
    document_no: str | None = None,
    description: str | None = None,
    source_table: str | None = None,
    source_id: int | None = None,
) -> None:
    if account_id <= 0 or amount <= 0:
        return
    conn.execute(
        """
        INSERT INTO account_movements(
          company_id, account_kind, account_id, movement_date, movement_type, direction, amount,
          document_no, description, source_table, source_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_id,
            account_kind,
            account_id,
            movement_date or date.today().isoformat(),
            movement_type,
            direction,
            amount,
            document_no,
            description,
            source_table,
            source_id,
        ),
    )


def reset_operational_tables(conn: sqlite3.Connection, company_id: int) -> None:
    for table in [
        "employee_overtime_entries",
        "employee_advances",
        "account_movements",
        "entity_attachments",
        "bank_transfer_vouchers",
        "bank_statement_lines",
        "purchase_invoices",
        "payment_instruments",
        "employees",
        "reference_values",
        "business_partners",
        "project_sites",
    ]:
        conn.execute(f"DELETE FROM {table} WHERE company_id = ?", (company_id,))


def preserved_employee_overrides(conn: sqlite3.Connection, company_id: int) -> dict:
    rows = conn.execute(
        """
        SELECT e.national_id_masked, e.full_name, e.monthly_salary, e.advance_amount,
               COALESCE(s.name, '') AS project_site
        FROM employees e
        LEFT JOIN project_sites s ON s.id = e.project_site_id
        WHERE e.company_id = ?
          AND (
            COALESCE(e.monthly_salary, 0) <> 0
            OR COALESCE(e.advance_amount, 0) <> 0
            OR COALESCE(s.name, '') <> ''
          )
        """
        ,
        (company_id,),
    ).fetchall()
    preserved = {}
    for row in rows:
        item = dict(row)
        for key in [item.get("national_id_masked"), item.get("full_name")]:
            if key:
                preserved[key] = item
    return preserved


def restore_employee_overrides(conn: sqlite3.Connection, overrides: dict, company_id: int) -> int:
    restored = 0
    if not overrides:
        return restored
    rows = conn.execute("SELECT id, national_id_masked, full_name FROM employees WHERE company_id = ?", (company_id,)).fetchall()
    for row in rows:
        item = overrides.get(row["national_id_masked"]) or overrides.get(row["full_name"])
        if not item:
            continue
        site_id = get_or_create_site(conn, item.get("project_site", ""), company_id) if item.get("project_site") else None
        conn.execute(
            """
            UPDATE employees
            SET monthly_salary = ?, advance_amount = ?, project_site_id = ?
            WHERE id = ? AND company_id = ?
            """,
            (
                to_float(item.get("monthly_salary")),
                to_float(item.get("advance_amount")),
                site_id,
                row["id"],
                company_id,
            ),
        )
        restored += 1
    return restored


def import_workbook(path: Path, original_name: str, company_id: int, actor: str = "local-admin") -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    summary = {
        "fileName": original_name,
        "sheetCount": len(wb.sheetnames),
        "sheets": {},
        "records": {
            "bankLines": 0,
            "purchaseInvoices": 0,
            "employees": 0,
            "paymentInstruments": 0,
            "partners": 0,
        },
        "warnings": [],
    }
    backup_path = backup_database("pre_import")
    if backup_path:
        summary["backupCreated"] = True
        summary["backupFile"] = Path(backup_path).name

    with connect() as conn:
        employee_overrides = preserved_employee_overrides(conn, company_id)
        reset_operational_tables(conn, company_id)

        if "Tanimlar" in wb.sheetnames:
            ws = wb["Tanimlar"]
            headers = header_map(ws)
            pairs = [
                ("Ana Grup", "bank_group"),
                ("Alt Kategori", "bank_sub_category"),
                ("Durum", "status"),
                ("Ödeme Yöntemi", "payment_method"),
                ("Çek Türü", "instrument_type"),
                ("Fatura Belge Türü", "invoice_document_type"),
                ("Ödeme Durumu", "payment_status"),
                ("KDV Oranı", "vat_rate"),
                ("Tevkifat", "withholding"),
            ]
            for row in range(2, ws.max_row + 1):
                for excel_col, group in pairs:
                    item = normalize_text(value(ws, row, headers, excel_col))
                    if item:
                        conn.execute(
                            """
                            INSERT OR IGNORE INTO reference_values(company_id, reference_group, value)
                            VALUES (?, ?, ?)
                            """,
                            (company_id, group, item),
                        )

        if "Personel" in wb.sheetnames:
            ws = wb["Personel"]
            headers = header_map(ws)
            rows = 0
            for row in range(2, ws.max_row + 1):
                first = normalize_text(value(ws, row, headers, "Adı"))
                last = normalize_text(value(ws, row, headers, "Soyadı"))
                if not first and not last:
                    continue
                site_id = get_or_create_site(conn, normalize_text(value(ws, row, headers, "Şantiye")), company_id)
                full_name = normalize_text(value(ws, row, headers, "Ad Soyad")) or f"{first} {last}".strip()
                conn.execute(
                    """
                    INSERT INTO employees(
                      company_id,
                      national_id_masked, first_name, last_name, full_name, hire_date, leave_date,
                      job_code, worked_days, status, project_site_id, monthly_salary,
                      advance_amount, iban_masked, phone_masked
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_id,
                        mask_identifier(value(ws, row, headers, "T.C. Kimlik No")),
                        first,
                        last,
                        full_name,
                        to_iso(value(ws, row, headers, "İşe Giriş Tarihi")),
                        to_iso(value(ws, row, headers, "İşe Çıkış Tarihi")),
                        normalize_text(value(ws, row, headers, "SGK Meslek Kodu")),
                        int(to_float(value(ws, row, headers, "Çalışılan Gün"))),
                        normalize_status(value(ws, row, headers, "Durum")) or "active",
                        site_id,
                        to_float(value(ws, row, headers, "Aylık Maaş (₺)")),
                        to_float(value(ws, row, headers, "Avans (₺)")),
                        mask_identifier(value(ws, row, headers, "IBAN")),
                        mask_identifier(value(ws, row, headers, "Telefon")),
                    ),
                )
                rows += 1
            summary["records"]["employees"] = rows
            summary["sheets"]["Personel"] = {"importedRows": rows, "target": "employees"}
            restored = restore_employee_overrides(conn, employee_overrides, company_id)
            if restored:
                summary["sheets"]["Personel"]["restoredManualOverrides"] = restored

        if "Banka_Ekstresi" in wb.sheetnames:
            ws = wb["Banka_Ekstresi"]
            headers = header_map(ws)
            rows = 0
            for row in range(2, ws.max_row + 1):
                if not value(ws, row, headers, "Tarih"):
                    continue
                conn.execute(
                    """
                    INSERT INTO bank_statement_lines(
                      company_id,
                      transaction_date, transaction_type, description, transaction_group,
                      sub_category, debit_amount, credit_amount, balance_amount, direction, net_amount
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_id,
                        to_iso(value(ws, row, headers, "Tarih")),
                        normalize_text(value(ws, row, headers, "İşlem")),
                        normalize_text(value(ws, row, headers, "Açıklama")),
                        normalize_text(value(ws, row, headers, "Grup")),
                        normalize_text(value(ws, row, headers, "Alt Kategori")),
                        to_float(value(ws, row, headers, "Borç (₺)")),
                        to_float(value(ws, row, headers, "Alacak (₺)")),
                        to_float(value(ws, row, headers, "Bakiye (₺)")),
                        normalize_text(value(ws, row, headers, "B/A")),
                        to_float(value(ws, row, headers, "Net Tutar (₺)")),
                    ),
                )
                rows += 1
            summary["records"]["bankLines"] = rows
            summary["sheets"]["Banka_Ekstresi"] = {"importedRows": rows, "target": "bank_statement_lines"}

        if "Gunluk_Fatura_Hareketleri" in wb.sheetnames:
            ws = wb["Gunluk_Fatura_Hareketleri"]
            headers = header_map(ws)
            invoice_numbers = {}
            rows = 0
            for row in range(2, ws.max_row + 1):
                if not value(ws, row, headers, "Tarih"):
                    continue
                partner_id = get_or_create_partner(conn, normalize_text(value(ws, row, headers, "Cari")), "vendor", company_id)
                site_id = get_or_create_site(conn, normalize_text(value(ws, row, headers, "Şantiye")), company_id)
                invoice_no = normalize_text(value(ws, row, headers, "Fatura No"))
                if invoice_no:
                    invoice_numbers[invoice_no] = invoice_numbers.get(invoice_no, 0) + 1
                conn.execute(
                    """
                    INSERT INTO purchase_invoices(
                      company_id,
                      invoice_date, document_type, invoice_no, partner_id, description,
                      purchase_amount, sales_amount, vat_rate, withholding_code, project_site_id,
                      cost_category, payment_status, due_date, vat_amount, withholding_amount,
                      gross_total, paid_amount, remaining_amount, delay_status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_id,
                        to_iso(value(ws, row, headers, "Tarih")),
                        normalize_text(value(ws, row, headers, "Belge Türü")) or "Alış",
                        invoice_no,
                        partner_id,
                        normalize_text(value(ws, row, headers, "Açıklama")),
                        to_float(value(ws, row, headers, "Alış")),
                        to_float(value(ws, row, headers, "Satış")),
                        to_float(value(ws, row, headers, "KDV")),
                        normalize_text(value(ws, row, headers, "Tevkifat")),
                        site_id,
                        normalize_text(value(ws, row, headers, "Kategori")),
                        normalize_status(value(ws, row, headers, "Ödeme Durumu")),
                        to_iso(value(ws, row, headers, "Vade")),
                        to_float(value(ws, row, headers, "KDV Tutarı")),
                        to_float(value(ws, row, headers, "Tevkifat Tutarı")),
                        to_float(value(ws, row, headers, "Genel Toplam")),
                        to_float(value(ws, row, headers, "Ödenen Tutar")),
                        to_float(value(ws, row, headers, "Kalan Tutar")),
                        normalize_status(value(ws, row, headers, "Gecikme Durumu")),
                    ),
                )
                rows += 1
            duplicates = [key for key, count in invoice_numbers.items() if count > 1]
            if duplicates:
                summary["warnings"].append(f"{len(duplicates)} fatura numarası tekrarlı görünüyor.")
            summary["records"]["purchaseInvoices"] = rows
            summary["sheets"]["Gunluk_Fatura_Hareketleri"] = {
                "importedRows": rows,
                "target": "purchase_invoices",
                "duplicateInvoiceNoCount": len(duplicates),
            }

        if "Cek_Senet_Takibi" in wb.sheetnames:
            ws = wb["Cek_Senet_Takibi"]
            headers = header_map(ws)
            rows = 0
            for row in range(2, ws.max_row + 1):
                amount = to_float(value(ws, row, headers, "Tutar (₺)"))
                due = value(ws, row, headers, "Vade Tarihi")
                if not amount and not due:
                    continue
                partner_id = get_or_create_partner(conn, normalize_text(value(ws, row, headers, "Cari / Firma")), "vendor", company_id)
                conn.execute(
                    """
                    INSERT INTO payment_instruments(
                      company_id,
                      instrument_type, partner_id, instrument_no, bank_name, issue_date,
                      due_date, amount, status, settlement_date, note
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_id,
                        normalize_text(value(ws, row, headers, "Tür")),
                        partner_id,
                        normalize_text(value(ws, row, headers, "Çek / Senet No")),
                        normalize_text(value(ws, row, headers, "Banka")),
                        to_iso(value(ws, row, headers, "Keşide Tarihi")),
                        to_iso(due),
                        amount,
                        normalize_status(value(ws, row, headers, "Durum")),
                        to_iso(value(ws, row, headers, "Tahsil / Ödeme Tarihi")),
                        normalize_text(value(ws, row, headers, "Not")),
                    ),
                )
                rows += 1
            summary["records"]["paymentInstruments"] = rows
            summary["sheets"]["Cek_Senet_Takibi"] = {"importedRows": rows, "target": "payment_instruments"}

        partner_count = conn.execute("SELECT COUNT(*) AS count FROM business_partners WHERE company_id = ?", (company_id,)).fetchone()["count"]
        site_count = conn.execute("SELECT COUNT(*) AS count FROM project_sites WHERE company_id = ?", (company_id,)).fetchone()["count"]
        summary["records"]["partners"] = partner_count
        summary["records"]["projectSites"] = site_count

        conn.execute(
            """
            INSERT INTO import_batches(company_id, file_name, sheet_count, summary_json)
            VALUES (?, ?, ?, ?)
            """,
            (company_id, original_name, len(wb.sheetnames), json.dumps(summary, ensure_ascii=False)),
        )
        audit_event(conn, company_id, actor, "import_workbook", "import_batches", None, None, summary)

    return summary


def dashboard_payload(user: dict | None = None, company_id: int | None = None) -> dict:
    init_db()
    with connect() as conn:
        if user:
            company_id = resolve_company_id(conn, user, company_id)
        else:
            company_id = company_id or ensure_default_company(conn)
        current_period = date.today().strftime("%Y-%m")

        def scalar(sql: str, params=()):
            row = conn.execute(sql, params).fetchone()
            return list(row)[0] if row else 0

        invoice_total = scalar("SELECT COALESCE(SUM(gross_total), 0) FROM purchase_invoices WHERE company_id = ?", (company_id,))
        invoice_remaining = scalar("SELECT COALESCE(SUM(remaining_amount), 0) FROM purchase_invoices WHERE company_id = ?", (company_id,))
        bank_net = scalar("SELECT COALESCE(SUM(net_amount), 0) FROM bank_statement_lines WHERE company_id = ?", (company_id,))
        instrument_total = scalar("SELECT COALESCE(SUM(amount), 0) FROM payment_instruments WHERE company_id = ?", (company_id,))
        duplicate_invoices = scalar(
            """
            SELECT COUNT(*) FROM (
              SELECT invoice_no
              FROM purchase_invoices
              WHERE company_id = ? AND invoice_no IS NOT NULL AND invoice_no <> ''
              GROUP BY invoice_no
              HAVING COUNT(*) > 1
            )
            """,
            (company_id,),
        )
        missing_due_dates = scalar(
            "SELECT COUNT(*) FROM purchase_invoices WHERE company_id = ? AND remaining_amount > 0 AND (due_date IS NULL OR due_date = '')",
            (company_id,),
        )
        missing_cost_category = scalar(
            "SELECT COUNT(*) FROM purchase_invoices WHERE company_id = ? AND (cost_category IS NULL OR cost_category = '')",
            (company_id,),
        )
        duplicate_employees = scalar(
            """
            SELECT COUNT(*)
            FROM (
              SELECT normalize_name, COUNT(*) AS count
              FROM (
                SELECT lower(trim(full_name)) AS normalize_name
                FROM employees
                WHERE company_id = ? AND full_name IS NOT NULL AND full_name <> '' AND COALESCE(status, 'active') <> 'deleted'
              )
              GROUP BY normalize_name
              HAVING COUNT(*) > 1
            )
            """,
            (company_id,),
        )
        unmatched_bank_lines = scalar("SELECT COUNT(*) FROM bank_statement_lines WHERE company_id = ? AND COALESCE(match_status, 'unmatched') <> 'matched'", (company_id,))
        pending_invoices = scalar("SELECT COUNT(*) FROM purchase_invoices WHERE company_id = ? AND remaining_amount > 0", (company_id,))

        latest_import = conn.execute(
            "SELECT file_name, imported_at, summary_json FROM import_batches WHERE company_id = ? ORDER BY id DESC LIMIT 1",
            (company_id,),
        ).fetchone()

        invoice_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT p.invoice_date, p.invoice_no, b.name AS partner, p.gross_total,
                       p.remaining_amount, p.payment_status, p.delay_status
                FROM purchase_invoices p
                LEFT JOIN business_partners b ON b.id = p.partner_id
                WHERE p.company_id = ?
                ORDER BY p.invoice_date DESC
                LIMIT 8
                """,
                (company_id,),
            )
        ]

        bank_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT id, transaction_date, transaction_type, transaction_group, sub_category, net_amount,
                       COALESCE(match_status, 'unmatched') AS match_status, match_type, account_code, match_note
                FROM bank_statement_lines
                WHERE company_id = ?
                ORDER BY CASE WHEN COALESCE(match_status, 'unmatched') = 'matched' THEN 1 ELSE 0 END,
                         transaction_date DESC
                """,
                (company_id,),
            )
        ]

        payable_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT p.id, p.invoice_date, p.due_date, p.invoice_no, b.name AS partner,
                       p.purchase_amount, p.vat_amount, p.gross_total, p.paid_amount,
                       p.remaining_amount, p.payment_status,
                       COALESCE(s.name, 'Atanmadı') AS project_site
                FROM purchase_invoices p
                LEFT JOIN business_partners b ON b.id = p.partner_id
                LEFT JOIN project_sites s ON s.id = p.project_site_id
                WHERE p.company_id = ?
                ORDER BY p.remaining_amount DESC, p.invoice_date DESC
                """,
                (company_id,),
            )
        ]

        partner_rows = []
        for row in conn.execute(
            """
            SELECT b.id, b.name, b.partner_type,
                   COALESCE(inv.invoice_count, 0) AS invoice_count,
                   COALESCE(inv.gross_total, 0) AS gross_total,
                   COALESCE(m.debit_total, 0) AS debit_total,
                   COALESCE(m.credit_total, 0) AS credit_total,
                   COALESCE(a.file_count, 0) AS attachment_count
            FROM business_partners b
            LEFT JOIN (
              SELECT partner_id, COUNT(*) AS invoice_count, COALESCE(SUM(gross_total), 0) AS gross_total
              FROM purchase_invoices
              WHERE company_id = ?
              GROUP BY partner_id
            ) inv ON inv.partner_id = b.id
            LEFT JOIN (
              SELECT account_id,
                     COALESCE(SUM(CASE WHEN direction = 'debit' THEN amount ELSE 0 END), 0) AS debit_total,
                     COALESCE(SUM(CASE WHEN direction = 'credit' THEN amount ELSE 0 END), 0) AS credit_total
              FROM account_movements
              WHERE company_id = ? AND account_kind = 'partner'
              GROUP BY account_id
            ) m ON m.account_id = b.id
            LEFT JOIN (
              SELECT entity_id, COUNT(*) AS file_count
              FROM entity_attachments
              WHERE company_id = ? AND entity_type = 'partner'
              GROUP BY entity_id
            ) a ON a.entity_id = b.id
            WHERE b.company_id = ?
            GROUP BY b.id
            ORDER BY ABS(credit_total - debit_total) DESC, gross_total DESC
            """,
            (company_id, company_id, company_id, company_id),
        ):
            item = dict(row)
            item["open_balance"] = max(float(item["credit_total"] or 0) - float(item["debit_total"] or 0), 0)
            partner_rows.append(item)

        employee_rows = []
        for row in conn.execute(
            """
            SELECT e.id, e.full_name, e.job_code, e.hire_date, e.leave_date, e.worked_days, e.status,
                   e.monthly_salary, e.advance_amount, e.iban_masked, COALESCE(s.name, '') AS project_site,
                   COALESCE(a.file_count, 0) AS attachment_count,
                   COALESCE(adv.advance_total, COALESCE(e.advance_amount, 0), 0) AS period_advance_total,
                   COALESCE(adv.advance_count, 0) AS advance_count,
                   COALESCE(ot.overtime_total, 0) AS overtime_total,
                   COALESCE(ot.overtime_hours, 0) AS overtime_hours,
                   COALESCE(ot.overtime_count, 0) AS overtime_count
            FROM employees e
            LEFT JOIN project_sites s ON s.id = e.project_site_id
            LEFT JOIN (
              SELECT entity_id, COUNT(*) AS file_count
              FROM entity_attachments
              WHERE company_id = ? AND entity_type = 'employee'
              GROUP BY entity_id
            ) a ON a.entity_id = e.id
            LEFT JOIN (
              SELECT employee_id, COUNT(*) AS advance_count, COALESCE(SUM(amount), 0) AS advance_total
              FROM employee_advances
              WHERE company_id = ? AND period = ?
              GROUP BY employee_id
            ) adv ON adv.employee_id = e.id
            LEFT JOIN (
              SELECT employee_id, COUNT(*) AS overtime_count,
                     COALESCE(SUM(hours), 0) AS overtime_hours,
                     COALESCE(SUM(amount), 0) AS overtime_total
              FROM employee_overtime_entries
              WHERE company_id = ? AND period = ?
              GROUP BY employee_id
            ) ot ON ot.employee_id = e.id
            WHERE e.company_id = ? AND COALESCE(e.status, 'active') <> 'deleted'
            ORDER BY e.full_name
            """,
            (company_id, company_id, current_period, company_id, current_period, company_id),
        ):
            item = dict(row)
            base_salary, payroll_days = payroll_base_salary(
                float(item.get("monthly_salary") or 0),
                item.get("hire_date"),
                item.get("leave_date"),
                current_period,
            )
            item["payroll_period"] = current_period
            item["payroll_days"] = payroll_days
            item["base_salary"] = base_salary
            item["paid_salary"] = max(
                base_salary + float(item.get("overtime_total") or 0) - float(item.get("period_advance_total") or 0),
                0,
            )
            item["advance_amount"] = float(item.get("period_advance_total") or 0)
            item["seniority_days"] = 0
            item["seniority_label"] = ""
            item["job_title"] = clean_job_title(item.get("job_code", ""))
            try:
                hired = datetime.fromisoformat(str(item["hire_date"])[:19]).date()
                days = max((date.today() - hired).days + 1, 0)
                item["seniority_days"] = days
                if days < 30:
                    item["seniority_label"] = f"{days} gün"
                else:
                    item["seniority_label"] = f"{days // 30} ay {days % 30} gün"
            except Exception:
                pass
            employee_rows.append(item)

        instrument_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT i.id, i.instrument_type, b.name AS partner, i.instrument_no, i.bank_name,
                       i.due_date, i.amount, i.status
                FROM payment_instruments i
                LEFT JOIN business_partners b ON b.id = i.partner_id
                WHERE i.company_id = ?
                ORDER BY i.due_date, i.amount DESC
                """,
                (company_id,),
            )
        ]

        vat_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT substr(invoice_date, 1, 7) AS period,
                       COALESCE(SUM(purchase_amount), 0) AS purchase_base,
                       COALESCE(SUM(vat_amount), 0) AS purchase_vat,
                       COALESCE(SUM(sales_amount), 0) AS sales_base,
                       COALESCE(SUM(withholding_amount), 0) AS withholding,
                       COALESCE(SUM(vat_amount - withholding_amount), 0) AS net_vat
                FROM purchase_invoices
                WHERE company_id = ? AND invoice_date IS NOT NULL AND invoice_date <> ''
                GROUP BY substr(invoice_date, 1, 7)
                ORDER BY period DESC
                """,
                (company_id,),
            )
        ]

        bank_group_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT COALESCE(transaction_group, 'Sınıfsız') AS name,
                       COUNT(*) AS line_count,
                       COALESCE(SUM(debit_amount), 0) AS debit_total,
                       COALESCE(SUM(credit_amount), 0) AS credit_total,
                       COALESCE(SUM(net_amount), 0) AS net_total
                FROM bank_statement_lines
                WHERE company_id = ?
                GROUP BY COALESCE(transaction_group, 'Sınıfsız')
                ORDER BY ABS(net_total) DESC, line_count DESC
                """,
                (company_id,),
            )
        ]

        project_site_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT id, name, status
                FROM project_sites
                WHERE company_id = ?
                ORDER BY name
                """,
                (company_id,),
            )
        ]

        movement_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT id, account_kind, account_id, movement_date, movement_type, direction,
                       amount, document_no, description, source_table, source_id
                FROM account_movements
                WHERE company_id = ?
                ORDER BY movement_date DESC, id DESC
                """,
                (company_id,),
            )
        ]

        attachment_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT id, entity_type, entity_id, file_name, mime_type, file_size, uploaded_at
                FROM entity_attachments
                WHERE company_id = ?
                ORDER BY uploaded_at DESC, id DESC
                """,
                (company_id,),
            )
        ]

        transfer_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT id, transfer_date, from_account_code, to_account_code, amount, description, source_bank_line_id
                FROM bank_transfer_vouchers
                WHERE company_id = ?
                ORDER BY transfer_date DESC, id DESC
                """,
                (company_id,),
            )
        ]

        employee_advance_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT a.id, a.employee_id, e.full_name, a.advance_date, a.period, a.amount, a.note
                FROM employee_advances a
                LEFT JOIN employees e ON e.id = a.employee_id
                WHERE a.company_id = ?
                ORDER BY a.advance_date DESC, a.id DESC
                """,
                (company_id,),
            )
        ]

        employee_overtime_rows = [
            dict(row)
            for row in conn.execute(
                """
                SELECT o.id, o.employee_id, e.full_name, o.overtime_date, o.period,
                       o.hours, o.hourly_rate, o.amount, o.note
                FROM employee_overtime_entries o
                LEFT JOIN employees e ON e.id = o.employee_id
                WHERE o.company_id = ?
                ORDER BY o.overtime_date DESC, o.id DESC
                """,
                (company_id,),
            )
        ]

        audit_rows = []
        if can_view_logs(user):
            audit_rows = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT a.id, a.actor, a.action, a.entity_name, a.entity_id,
                           a.old_value, a.new_value, a.created_at, c.name AS company_name
                    FROM audit_events a
                    LEFT JOIN companies c ON c.id = a.company_id
                    WHERE a.company_id = ?
                    ORDER BY a.created_at DESC, a.id DESC
                    LIMIT 250
                    """,
                    (company_id,),
                )
            ]

        selected_company = conn.execute("SELECT id, name, tax_number, status FROM companies WHERE id = ?", (company_id,)).fetchone()
        open_invoice_sum = scalar("SELECT COALESCE(SUM(remaining_amount), 0) FROM purchase_invoices WHERE company_id = ? AND remaining_amount > 0", (company_id,))
        due_instrument_sum = scalar("SELECT COALESCE(SUM(amount), 0) FROM payment_instruments WHERE company_id = ?", (company_id,))

        return {
            "kpis": {
                "invoiceCount": scalar("SELECT COUNT(*) FROM purchase_invoices WHERE company_id = ?", (company_id,)),
                "invoiceTotal": invoice_total,
                "invoiceRemaining": invoice_remaining,
                "bankLineCount": scalar("SELECT COUNT(*) FROM bank_statement_lines WHERE company_id = ?", (company_id,)),
                "bankNet": bank_net,
                "employeeCount": scalar("SELECT COUNT(*) FROM employees WHERE company_id = ? AND COALESCE(status, 'active') <> 'deleted'", (company_id,)),
                "paymentInstrumentCount": scalar("SELECT COUNT(*) FROM payment_instruments WHERE company_id = ?", (company_id,)),
                "paymentInstrumentTotal": instrument_total,
                "partnerCount": scalar("SELECT COUNT(*) FROM business_partners WHERE company_id = ?", (company_id,)),
                "pendingInvoices": pending_invoices,
                "duplicateInvoices": duplicate_invoices,
                "missingDueDates": missing_due_dates,
                "missingCostCategory": missing_cost_category,
                "duplicateEmployees": duplicate_employees,
                "unmatchedBankLines": unmatched_bank_lines,
            },
            "workQueue": [
                {
                    "label": "Ödeme bekleyen faturalar",
                    "count": pending_invoices,
                    "amount": open_invoice_sum,
                    "severity": "high" if open_invoice_sum else "normal",
                    "target": "ap",
                },
                {
                    "label": "Mutabakat bekleyen banka satırları",
                    "count": unmatched_bank_lines,
                    "amount": bank_net,
                    "severity": "normal",
                    "target": "bank",
                },
                {
                    "label": "Vadesi net olmayan açık faturalar",
                    "count": missing_due_dates,
                    "amount": 0,
                    "severity": "high" if missing_due_dates else "normal",
                    "target": "controls",
                },
                {
                    "label": "Çek / senet portföyü",
                    "count": scalar("SELECT COUNT(*) FROM payment_instruments WHERE company_id = ?", (company_id,)),
                    "amount": due_instrument_sum,
                    "severity": "medium",
                    "target": "payments",
                },
            ],
            "controls": [
                {
                    "name": "Tekrarlı fatura numarası",
                    "count": duplicate_invoices,
                    "owner": "Muhasebe",
                    "action": "Fatura kayıtlarını belge no ve cari bazında doğrula.",
                },
                {
                    "name": "Eksik vade",
                    "count": missing_due_dates,
                    "owner": "Finans",
                    "action": "Açık faturalar için vade zorunlu hale getir.",
                },
                {
                    "name": "Eksik maliyet kategorisi",
                    "count": missing_cost_category,
                    "owner": "Maliyet kontrol",
                    "action": "Gider kategorisini fatura girişinde zorunlu yap.",
                },
                {
                    "name": "Maaş bilgisi eksik personel",
                    "count": scalar("SELECT COUNT(*) FROM employees WHERE company_id = ? AND COALESCE(status, 'active') <> 'deleted' AND (monthly_salary IS NULL OR monthly_salary = 0)", (company_id,)),
                    "owner": "Muhasebe",
                    "action": "Personel maaş kartlarını tamamla; avans düşümü ödenecek maaşa otomatik yansır.",
                },
                {
                    "name": "Tekrarlı personel kartı",
                    "count": duplicate_employees,
                    "owner": "Muhasebe",
                    "action": "Aynı personelin iki kez maaş/avans takibine girmediğini doğrula.",
                },
            ],
            "latestImport": dict(latest_import) if latest_import else None,
            "recentInvoices": invoice_rows,
            "recentBankLines": bank_rows,
            "payables": payable_rows,
            "partners": partner_rows,
            "employees": employee_rows,
            "paymentInstruments": instrument_rows,
            "vatSummary": vat_rows,
            "bankGroups": bank_group_rows,
            "projectSites": project_site_rows,
            "accountMovements": movement_rows,
            "attachments": attachment_rows,
            "transferVouchers": transfer_rows,
            "employeeAdvances": employee_advance_rows,
            "employeeOvertime": employee_overtime_rows,
            "auditLogs": audit_rows,
            "companies": list_companies(conn) if user and user.get("role") == ADMIN_ROLE else ([dict(selected_company)] if selected_company else []),
            "adminUsers": list_users(conn) if user and user.get("role") == ADMIN_ROLE else [],
            "selectedCompany": dict(selected_company) if selected_company else None,
            "permissions": {
                "role": user.get("role") if user else ADMIN_ROLE,
                "canViewLogs": can_view_logs(user),
                "canSwitchCompany": bool(user and user.get("role") == ADMIN_ROLE),
                "canManageUsers": bool(user and user.get("role") == ADMIN_ROLE),
            },
            "reports": {
                "partnerDebit": sum(float(item.get("debit_total") or 0) for item in partner_rows),
                "partnerCredit": sum(float(item.get("credit_total") or 0) for item in partner_rows),
                "partnerOpenBalance": sum(float(item.get("open_balance") or 0) for item in partner_rows),
                "employeeNetPayable": sum(float(item.get("paid_salary") or 0) for item in employee_rows),
                "unmatchedBankLines": unmatched_bank_lines,
                "attachmentCount": len(attachment_rows),
            },
        }


class ERPHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PUBLIC), **kwargs)

    def end_headers(self) -> None:
        static_path = urlparse(self.path).path
        if static_path.endswith((".html", ".js", ".css")) or static_path in {"/", "/login"}:
            self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header(
            "Content-Security-Policy",
            "default-src 'self'; script-src 'self'; style-src 'self'; img-src 'self' data:; "
            "connect-src 'self'; base-uri 'none'; frame-ancestors 'none'; form-action 'self'",
        )
        super().end_headers()

    def current_user(self) -> dict | None:
        return current_user_from_cookie(self.headers.get("Cookie"))

    def send_redirect(self, location: str) -> None:
        self.send_response(302)
        self.send_header("Location", location)
        self.end_headers()

    def send_auth_cookie(self, token: str) -> None:
        parts = [
            f"{SESSION_COOKIE}={token}",
            "Path=/",
            "HttpOnly",
            "SameSite=Lax",
            f"Max-Age={SESSION_DAYS * 24 * 60 * 60}",
        ]
        if should_send_secure_cookie(self.headers):
            parts.append("Secure")
        self.send_header("Set-Cookie", "; ".join(parts))

    def clear_auth_cookie(self) -> None:
        parts = [f"{SESSION_COOKIE}=", "Path=/", "HttpOnly", "SameSite=Lax", "Max-Age=0"]
        if should_send_secure_cookie(self.headers):
            parts.append("Secure")
        self.send_header("Set-Cookie", "; ".join(parts))

    def send_json(self, payload: dict, status: int = 200) -> None:
        encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def selected_company_id(self, user: dict, requested=None) -> int:
        with connect() as conn:
            return resolve_company_id(conn, user, requested)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/login", "/login.html"):
            if self.current_user():
                self.send_redirect("/")
                return
            self.path = "/login.html"
            return super().do_GET()
        if parsed.path == "/api/auth/config":
            with connect() as conn:
                self.send_json(
                    {
                        "registrationAllowed": registration_allowed(conn),
                        "firstUser": user_count(conn) == 0,
                        "inviteRequired": user_count(conn) > 0 and bool(INVITE_CODE) and not ALLOW_OPEN_REGISTRATION,
                    }
                )
            return
        if parsed.path == "/api/me":
            user = self.current_user()
            if not user:
                self.send_json({"authenticated": False}, 401)
                return
            with connect() as conn:
                try:
                    company_id = resolve_company_id(conn, user, parse_qs(parsed.query).get("companyId", [None])[0])
                    selected_company = conn.execute("SELECT id, name, tax_number, status FROM companies WHERE id = ?", (company_id,)).fetchone()
                    companies = list_companies(conn) if user.get("role") == ADMIN_ROLE else ([dict(selected_company)] if selected_company else [])
                except CompanyAccessError:
                    company_id = None
                    selected_company = None
                    companies = list_companies(conn) if user.get("role") == ADMIN_ROLE else []
            self.send_json(
                {
                    "authenticated": True,
                    "user": user,
                    "companies": companies,
                    "selectedCompany": dict(selected_company) if selected_company else None,
                    "companyAssigned": company_id is not None,
                    "permissions": {
                        "role": user.get("role"),
                        "canViewLogs": can_view_logs(user),
                        "canSwitchCompany": user.get("role") == ADMIN_ROLE,
                        "canManageUsers": user.get("role") == ADMIN_ROLE,
                    },
                }
            )
            return
        if parsed.path == "/api/dashboard":
            user = self.current_user()
            if not user:
                self.send_json({"error": "Oturum gerekli."}, 401)
                return
            requested_company_id = parse_qs(parsed.query).get("companyId", [None])[0] or company_id_from_request(self)
            try:
                self.send_json(dashboard_payload(user, requested_company_id))
            except CompanyAccessError as exc:
                self.send_json({"error": str(exc), "companyAssigned": False}, 403)
            return
        if parsed.path == "/api/health":
            self.send_json({"status": "ok"})
            return
        if parsed.path.startswith("/api/attachments/"):
            user = self.current_user()
            if not user:
                self.send_json({"error": "Oturum gerekli."}, 401)
                return
            try:
                attachment_id = int(parsed.path.rsplit("/", 1)[-1])
            except ValueError:
                self.send_json({"error": "Geçerli dosya seçilmedi."}, 400)
                return
            with connect() as conn:
                try:
                    company_id = resolve_company_id(conn, user, parse_qs(parsed.query).get("companyId", [None])[0])
                except CompanyAccessError as exc:
                    self.send_json({"error": str(exc)}, 403)
                    return
                row = conn.execute("SELECT * FROM entity_attachments WHERE id = ? AND company_id = ?", (attachment_id, company_id)).fetchone()
            if not row:
                self.send_json({"error": "Dosya bulunamadı."}, 404)
                return
            path = ATTACHMENTS / row["stored_name"]
            if not path.exists() or path.parent != ATTACHMENTS:
                self.send_json({"error": "Dosya bulunamadı."}, 404)
                return
            payload = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", row["mime_type"] or "application/pdf")
            self.send_header("Content-Length", str(len(payload)))
            self.send_header("Content-Disposition", f'inline; filename="{Path(row["file_name"]).name}"')
            self.end_headers()
            self.wfile.write(payload)
            return
        if parsed.path in ("/", "/index.html"):
            if not self.current_user():
                self.send_redirect("/login")
                return
            self.path = "/index.html"
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if not same_origin_allowed(self.headers):
            self.send_json({"error": "İstek kaynağı reddedildi."}, 403)
            return

        if parsed.path == "/api/register":
            try:
                payload = read_json_body(self)
                email = normalize_text(payload.get("email")).casefold()
                if rate_limited(f"register:{self.client_address[0]}:{email}", 5, 15 * 60):
                    self.send_json({"error": "Çok fazla kayıt denemesi. Bir süre sonra tekrar dene."}, 429)
                    return
                full_name = normalize_text(payload.get("fullName"))
                password = str(payload.get("password") or "")
                invite_code = str(payload.get("inviteCode") or "")
                if "@" not in email or "." not in email:
                    self.send_json({"error": "Geçerli e-posta gir."}, 400)
                    return
                if not full_name:
                    self.send_json({"error": "Ad soyad zorunlu."}, 400)
                    return
                if len(password) < 10:
                    self.send_json({"error": "Parola en az 10 karakter olmalı."}, 400)
                    return
                with connect() as conn:
                    default_company_id = ensure_default_company(conn)
                    count = user_count(conn)
                    if count > 0 and not ALLOW_OPEN_REGISTRATION:
                        if not INVITE_CODE or not hmac.compare_digest(invite_code, INVITE_CODE):
                            self.send_json({"error": "Kayıt için geçerli davet kodu gerekli."}, 403)
                            return
                    password_hash, salt, iterations = hash_password(password)
                    role = "admin" if count == 0 else "accountant"
                    assigned_company_id = default_company_id if role == ADMIN_ROLE else None
                    cur = conn.execute(
                        """
                        INSERT INTO users(company_id, email, full_name, role, password_hash, password_salt, password_iterations)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (assigned_company_id, email, full_name, role, password_hash, salt, iterations),
                    )
                    token = create_session(
                        conn,
                        int(cur.lastrowid),
                        self.headers.get("User-Agent", ""),
                        self.client_address[0],
                    )
                    audit_event(conn, assigned_company_id or default_company_id, email, "register_user", "users", cur.lastrowid, None, {"role": role, "company_id": assigned_company_id})
                body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.send_auth_cookie(token)
                self.end_headers()
                self.wfile.write(body)
            except sqlite3.IntegrityError:
                self.send_json({"error": "Bu e-posta zaten kayıtlı."}, 409)
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Kayıt oluşturulamadı: {exc}"}, 500)
            return

        if parsed.path == "/api/login":
            try:
                payload = read_json_body(self)
                email = normalize_text(payload.get("email")).casefold()
                password = str(payload.get("password") or "")
                if rate_limited(f"login:{self.client_address[0]}:{email}", 10, 15 * 60):
                    self.send_json({"error": "Çok fazla giriş denemesi. Bir süre sonra tekrar dene."}, 429)
                    return
                with connect() as conn:
                    row = conn.execute(
                        """
                        SELECT id, company_id, email, full_name, role, password_hash, password_salt, password_iterations
                        FROM users
                        WHERE email = ? AND is_active = 1
                        """,
                        (email,),
                    ).fetchone()
                    if not row or not verify_password(password, row["password_hash"], row["password_salt"], int(row["password_iterations"])):
                        self.send_json({"error": "E-posta veya parola hatalı."}, 401)
                        return
                    token = create_session(
                        conn,
                        int(row["id"]),
                        self.headers.get("User-Agent", ""),
                        self.client_address[0],
                    )
                    conn.execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP WHERE id = ?", (row["id"],))
                body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.send_auth_cookie(token)
                self.end_headers()
                self.wfile.write(body)
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Giriş yapılamadı: {exc}"}, 500)
            return

        if parsed.path == "/api/logout":
            jar = cookies.SimpleCookie()
            try:
                jar.load(self.headers.get("Cookie", ""))
            except cookies.CookieError:
                pass
            morsel = jar.get(SESSION_COOKIE)
            if morsel:
                with connect() as conn:
                    conn.execute("DELETE FROM auth_sessions WHERE token_hash = ?", (hash_token(morsel.value),))
            body = b'{"ok":true}'
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.clear_auth_cookie()
            self.end_headers()
            self.wfile.write(body)
            return

        user = self.current_user() if parsed.path.startswith("/api/") else None
        if parsed.path.startswith("/api/") and not user:
            self.send_json({"error": "Oturum gerekli."}, 401)
            return
        if parsed.path.startswith("/api/") and user.get("role") != ADMIN_ROLE and not parse_company_id(user.get("company_id")):
            self.send_json({"error": "Bu hesap henüz bir firmaya bağlanmamış. Admin ataması gerekli."}, 403)
            return

        if parsed.path == "/api/companies":
            try:
                if user.get("role") != ADMIN_ROLE:
                    self.send_json({"error": "Firma yönetimi için admin yetkisi gerekli."}, 403)
                    return
                payload = read_json_body(self)
                name = normalize_text(payload.get("name"))
                tax_number = normalize_text(payload.get("taxNumber")) or None
                if not name:
                    self.send_json({"error": "Firma adı zorunlu."}, 400)
                    return
                with connect() as conn:
                    cur = conn.execute(
                        "INSERT INTO companies(name, tax_number) VALUES (?, ?)",
                        (name, tax_number),
                    )
                    company_id = int(cur.lastrowid)
                    audit_event(conn, company_id, user["email"], "create_company", "companies", company_id, None, {"name": name, "tax_number": tax_number})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except sqlite3.IntegrityError:
                self.send_json({"error": "Bu firma adı zaten kayıtlı."}, 409)
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Firma oluşturulamadı: {exc}"}, 500)
            return

        if parsed.path == "/api/admin/companies/update":
            try:
                if user.get("role") != ADMIN_ROLE:
                    self.send_json({"error": "Firma yönetimi için admin yetkisi gerekli."}, 403)
                    return
                payload = read_json_body(self)
                company_id = parse_company_id(payload.get("companyId"))
                name = normalize_text(payload.get("name"))
                tax_number = normalize_text(payload.get("taxNumber")) or None
                status = normalize_text(payload.get("status")) or "active"
                if not company_id:
                    self.send_json({"error": "Geçerli firma seçilmedi."}, 400)
                    return
                if not name:
                    self.send_json({"error": "Firma adı zorunlu."}, 400)
                    return
                if status not in {"active", "archived"}:
                    self.send_json({"error": "Geçerli firma durumu seçilmedi."}, 400)
                    return
                with connect() as conn:
                    before = conn.execute(
                        "SELECT id, name, tax_number, status FROM companies WHERE id = ?",
                        (company_id,),
                    ).fetchone()
                    if not before:
                        self.send_json({"error": "Firma bulunamadı."}, 404)
                        return
                    conn.execute(
                        "UPDATE companies SET name = ?, tax_number = ?, status = ? WHERE id = ?",
                        (name, tax_number, status, company_id),
                    )
                    audit_event(
                        conn,
                        company_id,
                        user["email"],
                        "update_company",
                        "companies",
                        company_id,
                        dict(before),
                        {"name": name, "tax_number": tax_number, "status": status},
                    )
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except sqlite3.IntegrityError:
                self.send_json({"error": "Bu firma adı zaten kayıtlı."}, 409)
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Firma güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/admin/users/assign":
            try:
                if user.get("role") != ADMIN_ROLE:
                    self.send_json({"error": "Kullanıcı yönetimi için admin yetkisi gerekli."}, 403)
                    return
                payload = read_json_body(self)
                target_user_id = int(payload.get("userId", 0))
                full_name = normalize_text(payload.get("fullName"))
                email = normalize_text(payload.get("email")).casefold()
                role = normalize_text(payload.get("role")) or ACCOUNTANT_ROLE
                target_company_id = parse_company_id(payload.get("companyId"))
                is_active = 1 if bool(payload.get("isActive", True)) else 0
                if target_user_id <= 0:
                    self.send_json({"error": "Geçerli kullanıcı seçilmedi."}, 400)
                    return
                if not full_name:
                    self.send_json({"error": "Ad soyad zorunlu."}, 400)
                    return
                if "@" not in email or "." not in email:
                    self.send_json({"error": "Geçerli e-posta gir."}, 400)
                    return
                if role not in ASSIGNABLE_ROLES:
                    self.send_json({"error": "Geçerli rol seçilmedi."}, 400)
                    return
                if role in {OWNER_ROLE, ACCOUNTANT_ROLE} and not target_company_id:
                    self.send_json({"error": "Şirket sahibi ve muhasebeci için firma zorunlu."}, 400)
                    return
                if target_user_id == int(user["id"]) and (role != ADMIN_ROLE or not is_active):
                    self.send_json({"error": "Kendi admin hesabını pasifleştiremez veya admin rolünden çıkaramazsın."}, 400)
                    return
                with connect() as conn:
                    before = conn.execute(
                        "SELECT id, company_id, email, full_name, role, is_active FROM users WHERE id = ?",
                        (target_user_id,),
                    ).fetchone()
                    if not before:
                        self.send_json({"error": "Kullanıcı bulunamadı."}, 404)
                        return
                    target_company = conn.execute("SELECT id, status FROM companies WHERE id = ?", (target_company_id,)).fetchone() if target_company_id else None
                    if target_company_id and not target_company:
                        self.send_json({"error": "Firma bulunamadı."}, 404)
                        return
                    if role in {OWNER_ROLE, ACCOUNTANT_ROLE} and target_company and normalize_text(target_company["status"]) == "archived":
                        self.send_json({"error": "Arşivli firmaya şirket sahibi veya muhasebeci atanamaz."}, 400)
                        return
                    if role == ADMIN_ROLE and not target_company_id:
                        target_company_id = parse_company_id(before["company_id"]) or ensure_default_company(conn)
                    conn.execute(
                        """
                        UPDATE users
                        SET company_id = ?, email = ?, full_name = ?, role = ?, is_active = ?
                        WHERE id = ?
                        """,
                        (target_company_id, email, full_name, role, is_active, target_user_id),
                    )
                    audit_event(
                        conn,
                        target_company_id or parse_company_id(before["company_id"]) or ensure_default_company(conn),
                        user["email"],
                        "assign_user_access",
                        "users",
                        target_user_id,
                        dict(before),
                        {"company_id": target_company_id, "email": email, "full_name": full_name, "role": role, "is_active": is_active},
                    )
                self.send_json({"dashboard": dashboard_payload(user, target_company_id)})
            except sqlite3.IntegrityError:
                self.send_json({"error": "Bu e-posta başka bir kullanıcıda kayıtlı."}, 409)
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Kullanıcı güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/partners":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                name = normalize_text(payload.get("name"))
                partner_type = normalize_text(payload.get("partnerType")) or "vendor"
                if not name:
                    self.send_json({"error": "Cari adı zorunlu."}, 400)
                    return
                with connect() as conn:
                    partner_id = get_or_create_partner(conn, name, partner_type, company_id)
                    audit_event(conn, company_id, user["email"], "create_partner", "business_partners", partner_id, None, {"name": name, "partner_type": partner_type})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Cari kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                full_name = normalize_text(payload.get("fullName"))
                if not full_name:
                    self.send_json({"error": "Personel adı zorunlu."}, 400)
                    return
                parts = full_name.split()
                first_name = parts[0]
                last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
                monthly_salary = to_float(payload.get("monthlySalary"))
                advance_amount = to_float(payload.get("advanceAmount"))
                worked_days = int(to_float(payload.get("workedDays")))
                if monthly_salary < 0 or advance_amount < 0 or worked_days < 0:
                    self.send_json({"error": "Maaş, avans ve gün negatif olamaz."}, 400)
                    return
                with connect() as conn:
                    duplicate = conn.execute(
                        "SELECT id FROM employees WHERE company_id = ? AND lower(trim(full_name)) = lower(trim(?)) AND COALESCE(status, 'active') <> 'deleted' LIMIT 1",
                        (company_id, full_name),
                    ).fetchone()
                    if duplicate:
                        self.send_json({"error": "Bu ad soyadla kayıtlı personel var. Mevcut kartı kontrol et."}, 409)
                        return
                    site_id = get_or_create_site(conn, payload.get("projectSite"), company_id)
                    cur = conn.execute(
                        """
                        INSERT INTO employees(
                          company_id, first_name, last_name, full_name, hire_date, job_code,
                          leave_date, worked_days, project_site_id, monthly_salary, advance_amount, iban_masked, status
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
                        """,
                        (
                            company_id,
                            first_name,
                            last_name,
                            full_name,
                            normalize_text(payload.get("hireDate")) or None,
                            normalize_text(payload.get("jobTitle")) or None,
                            normalize_text(payload.get("leaveDate")) or None,
                            worked_days,
                            site_id,
                            monthly_salary,
                            advance_amount,
                            normalize_text(payload.get("iban")) or None,
                        ),
                    )
                    audit_event(conn, company_id, user["email"], "create_employee", "employees", cur.lastrowid, None, {"full_name": full_name})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Personel kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/purchase-invoices":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                partner_name = normalize_text(payload.get("partnerName"))
                if not partner_name:
                    self.send_json({"error": "Cari adı zorunlu."}, 400)
                    return
                gross_total = to_float(payload.get("grossTotal"))
                paid_amount = to_float(payload.get("paidAmount"))
                vat_rate = to_float(payload.get("vatRate"))
                if gross_total < 0 or paid_amount < 0 or vat_rate < 0:
                    self.send_json({"error": "Tutarlar negatif olamaz."}, 400)
                    return
                remaining_amount = max(gross_total - paid_amount, 0)
                payment_status = "paid" if remaining_amount == 0 else ("partial" if paid_amount > 0 else "pending")
                due_date = normalize_text(payload.get("dueDate")) or None
                if remaining_amount > 0 and not due_date:
                    self.send_json({"error": "Açık fatura için vade tarihi zorunlu."}, 400)
                    return
                if due_date and remaining_amount > 0:
                    try:
                        if datetime.fromisoformat(due_date[:10]).date() < date.today():
                            payment_status = "overdue"
                    except ValueError:
                        pass
                vat_amount = gross_total * vat_rate / (100 + vat_rate) if vat_rate else 0
                purchase_amount = gross_total - vat_amount
                with connect() as conn:
                    partner_id = get_or_create_partner(conn, partner_name, "vendor", company_id)
                    site_id = get_or_create_site(conn, payload.get("projectSite"), company_id)
                    cur = conn.execute(
                        """
                        INSERT INTO purchase_invoices(
                          company_id, invoice_date, invoice_no, partner_id, description, purchase_amount,
                          vat_rate, project_site_id, cost_category, payment_status, due_date,
                          vat_amount, gross_total, paid_amount, remaining_amount
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            company_id,
                            normalize_text(payload.get("invoiceDate")) or None,
                            normalize_text(payload.get("invoiceNo")) or None,
                            partner_id,
                            normalize_text(payload.get("description")) or None,
                            purchase_amount,
                            vat_rate,
                            site_id,
                            normalize_text(payload.get("costCategory")) or None,
                            payment_status,
                            due_date,
                            vat_amount,
                            gross_total,
                            paid_amount,
                            remaining_amount,
                        ),
                    )
                    invoice_id = int(cur.lastrowid)
                    add_account_movement(
                        conn,
                        company_id,
                        "partner",
                        int(partner_id or 0),
                        normalize_text(payload.get("invoiceDate")) or date.today().isoformat(),
                        "invoice",
                        "credit",
                        gross_total,
                        normalize_text(payload.get("invoiceNo")) or None,
                        normalize_text(payload.get("description")) or "Alış faturası",
                        "purchase_invoices",
                        invoice_id,
                    )
                    if paid_amount > 0:
                        add_account_movement(
                            conn,
                            company_id,
                            "partner",
                            int(partner_id or 0),
                            normalize_text(payload.get("invoiceDate")) or date.today().isoformat(),
                            "payment",
                            "debit",
                            paid_amount,
                            normalize_text(payload.get("invoiceNo")) or None,
                            "Fatura ödemesi",
                            "purchase_invoices",
                            invoice_id,
                        )
                    audit_event(conn, company_id, user["email"], "create_purchase_invoice", "purchase_invoices", invoice_id, None, {"partner": partner_name, "gross_total": gross_total})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Fatura kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/account-movements":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                account_kind = normalize_text(payload.get("accountKind"))
                account_id = int(payload.get("accountId", 0))
                movement_type = normalize_text(payload.get("movementType")) or "debit_note"
                direction = normalize_text(payload.get("direction")) or "debit"
                amount = to_float(payload.get("amount"))
                movement_date = normalize_text(payload.get("movementDate")) or date.today().isoformat()
                document_no = normalize_text(payload.get("documentNo")) or None
                description = normalize_text(payload.get("description")) or None
                if account_kind not in {"partner", "employee"}:
                    self.send_json({"error": "Geçerli kart tipi seç."}, 400)
                    return
                if direction not in {"debit", "credit"}:
                    self.send_json({"error": "Borç/alacak yönü seç."}, 400)
                    return
                if account_id <= 0 or amount <= 0:
                    self.send_json({"error": "Kart ve tutar zorunlu."}, 400)
                    return
                with connect() as conn:
                    table = "business_partners" if account_kind == "partner" else "employees"
                    if not conn.execute(f"SELECT id FROM {table} WHERE id = ? AND company_id = ?", (account_id, company_id)).fetchone():
                        self.send_json({"error": "Kart bulunamadı."}, 404)
                        return
                    add_account_movement(
                        conn,
                        company_id,
                        account_kind,
                        account_id,
                        movement_date,
                        movement_type,
                        direction,
                        amount,
                        document_no,
                        description,
                        "manual",
                        None,
                    )
                    audit_event(conn, company_id, user["email"], "create_account_movement", "account_movements", account_id, None, payload)
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Hareket kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/bank/transfer":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                transfer_date = normalize_text(payload.get("transferDate")) or date.today().isoformat()
                from_account_code = normalize_text(payload.get("fromAccountCode"))
                to_account_code = normalize_text(payload.get("toAccountCode"))
                amount = to_float(payload.get("amount"))
                description = normalize_text(payload.get("description")) or None
                source_bank_line_id = int(to_float(payload.get("sourceBankLineId"))) if payload.get("sourceBankLineId") else None
                if not from_account_code or not to_account_code or amount <= 0:
                    self.send_json({"error": "Virman için çıkış hesabı, giriş hesabı ve tutar zorunlu."}, 400)
                    return
                with connect() as conn:
                    if source_bank_line_id and not conn.execute(
                        "SELECT id FROM bank_statement_lines WHERE id = ? AND company_id = ?",
                        (source_bank_line_id, company_id),
                    ).fetchone():
                        self.send_json({"error": "Ekstre satırı bulunamadı."}, 404)
                        return
                    cur = conn.execute(
                        """
                        INSERT INTO bank_transfer_vouchers(
                          company_id, transfer_date, from_account_code, to_account_code, amount, description, source_bank_line_id
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (company_id, transfer_date, from_account_code, to_account_code, amount, description, source_bank_line_id),
                    )
                    if source_bank_line_id:
                        conn.execute(
                            """
                            UPDATE bank_statement_lines
                            SET match_status = 'matched',
                                match_type = 'transfer',
                                account_code = ?,
                                match_note = ?,
                                matched_at = CURRENT_TIMESTAMP
                            WHERE id = ? AND company_id = ?
                            """,
                            (f"{from_account_code}>{to_account_code}", description, source_bank_line_id, company_id),
                        )
                    audit_event(conn, company_id, user["email"], "create_bank_transfer", "bank_transfer_vouchers", cur.lastrowid, None, payload)
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Virman fişi kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/purchase-invoices/mark-paid":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                ids = [int(item) for item in payload.get("ids", []) if int(item) > 0]
                if not ids:
                    self.send_json({"error": "Seçili fatura yok."}, 400)
                    return
                placeholders = ",".join("?" for _ in ids)
                with connect() as conn:
                    rows_before = [
                        dict(row)
                        for row in conn.execute(
                            f"""
                            SELECT id, partner_id, invoice_date, invoice_no, remaining_amount
                            FROM purchase_invoices
                            WHERE company_id = ? AND id IN ({placeholders})
                            """,
                            [company_id, *ids],
                        )
                    ]
                    conn.execute(
                        f"""
                        UPDATE purchase_invoices
                        SET paid_amount = gross_total,
                            remaining_amount = 0,
                            payment_status = 'paid'
                        WHERE company_id = ? AND id IN ({placeholders})
                        """,
                        [company_id, *ids],
                    )
                    for invoice in rows_before:
                        remaining = float(invoice.get("remaining_amount") or 0)
                        partner_id = int(invoice.get("partner_id") or 0)
                        if partner_id > 0 and remaining > 0:
                            add_account_movement(
                                conn,
                                company_id,
                                "partner",
                                partner_id,
                                date.today().isoformat(),
                                "payment",
                                "debit",
                                remaining,
                                invoice.get("invoice_no"),
                                "Ödendi işaretleme",
                                "purchase_invoices",
                                int(invoice["id"]),
                            )
                    audit_event(conn, company_id, user["email"], "bulk_mark_paid", "purchase_invoices", None, None, {"ids": ids})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Faturalar güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/bank/match":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                line_id = int(payload.get("lineId", 0))
                match_type = normalize_text(payload.get("matchType")) or "expense"
                account_code = normalize_text(payload.get("accountCode"))
                match_note = normalize_text(payload.get("matchNote"))
                partner_id = int(to_float(payload.get("partnerId"))) if payload.get("partnerId") else None
                invoice_id = int(to_float(payload.get("invoiceId"))) if payload.get("invoiceId") else None
                if line_id <= 0:
                    self.send_json({"error": "Geçerli banka satırı seçilmedi."}, 400)
                    return
                if match_type not in {"invoice", "partner", "expense", "transfer", "payroll"}:
                    self.send_json({"error": "Geçerli eşleştirme türü seç."}, 400)
                    return
                if match_type == "invoice" and not invoice_id:
                    self.send_json({"error": "Fatura eşleştirmesi için fatura seç."}, 400)
                    return
                if match_type in {"expense", "transfer", "payroll"} and not account_code:
                    self.send_json({"error": "Hesap kodu zorunlu."}, 400)
                    return
                with connect() as conn:
                    before = conn.execute("SELECT * FROM bank_statement_lines WHERE id = ? AND company_id = ?", (line_id, company_id)).fetchone()
                    if not before:
                        self.send_json({"error": "Banka satırı bulunamadı."}, 404)
                        return
                    if partner_id:
                        if not conn.execute("SELECT id FROM business_partners WHERE id = ? AND company_id = ?", (partner_id, company_id)).fetchone():
                            self.send_json({"error": "Cari bulunamadı."}, 404)
                            return
                    invoice_row = None
                    if invoice_id:
                        invoice_row = conn.execute("SELECT id, partner_id, invoice_no FROM purchase_invoices WHERE id = ? AND company_id = ?", (invoice_id, company_id)).fetchone()
                        if not invoice_row:
                            self.send_json({"error": "Fatura bulunamadı."}, 404)
                            return
                    conn.execute(
                        """
                        UPDATE bank_statement_lines
                        SET match_status = 'matched',
                            match_type = ?,
                            matched_partner_id = ?,
                            matched_invoice_id = ?,
                            account_code = ?,
                            match_note = ?,
                            matched_at = CURRENT_TIMESTAMP
                        WHERE id = ? AND company_id = ?
                        """,
                        (match_type, partner_id, invoice_id, account_code, match_note, line_id, company_id),
                    )
                    conn.execute(
                        "DELETE FROM account_movements WHERE company_id = ? AND source_table = 'bank_statement_lines' AND source_id = ?",
                        (company_id, line_id),
                    )
                    resolved_partner_id = partner_id or (int(invoice_row["partner_id"] or 0) if invoice_row else None)
                    net_amount = float(before["net_amount"] or 0)
                    if match_type in {"invoice", "partner"} and resolved_partner_id and net_amount:
                        add_account_movement(
                            conn,
                            company_id,
                            "partner",
                            int(resolved_partner_id),
                            normalize_text(before["transaction_date"]) or date.today().isoformat(),
                            "payment" if net_amount < 0 else "collection",
                            "debit" if net_amount < 0 else "credit",
                            abs(net_amount),
                            invoice_row["invoice_no"] if invoice_row else None,
                            match_note or normalize_text(before["transaction_type"]) or "Banka hareketi",
                            "bank_statement_lines",
                            line_id,
                        )
                    audit_event(
                        conn,
                        company_id,
                        user["email"],
                        "match_bank_line",
                        "bank_statement_lines",
                        line_id,
                        dict(before),
                        {
                            "match_type": match_type,
                            "partner_id": partner_id,
                            "invoice_id": invoice_id,
                            "account_code": account_code,
                            "match_note": match_note,
                        },
                    )
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Banka satırı eşleştirilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/bulk-site":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                ids = [int(item) for item in payload.get("ids", []) if int(item) > 0]
                site_name = normalize_text(payload.get("projectSiteName"))
                if not ids:
                    self.send_json({"error": "Seçili personel yok."}, 400)
                    return
                placeholders = ",".join("?" for _ in ids)
                with connect() as conn:
                    site_id = get_or_create_site(conn, site_name, company_id) if site_name else None
                    conn.execute(
                        f"UPDATE employees SET project_site_id = ? WHERE company_id = ? AND id IN ({placeholders})",
                        [site_id, company_id, *ids],
                    )
                    audit_event(conn, company_id, user["email"], "bulk_update_employee_site", "employees", None, None, {"ids": ids, "project_site": site_name})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Şantiye güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/bulk-advance":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                ids = [int(item) for item in payload.get("ids", []) if int(item) > 0]
                advance_amount = to_float(payload.get("advanceAmount"))
                advance_date = normalize_text(payload.get("advanceDate")) or date.today().isoformat()
                period = period_from_date(advance_date)
                note = normalize_text(payload.get("note")) or None
                if not ids:
                    self.send_json({"error": "Seçili personel yok."}, 400)
                    return
                if advance_amount < 0:
                    self.send_json({"error": "Avans negatif olamaz."}, 400)
                    return
                placeholders = ",".join("?" for _ in ids)
                with connect() as conn:
                    valid_ids = [
                        int(row["id"])
                        for row in conn.execute(
                            f"SELECT id FROM employees WHERE company_id = ? AND id IN ({placeholders})",
                            [company_id, *ids],
                        )
                    ]
                    for employee_id in valid_ids:
                        conn.execute(
                            """
                            INSERT INTO employee_advances(company_id, employee_id, advance_date, period, amount, note)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """,
                            (company_id, employee_id, advance_date, period, advance_amount, note),
                        )
                    if not valid_ids:
                        self.send_json({"error": "Personel bulunamadı."}, 404)
                        return
                    valid_placeholders = ",".join("?" for _ in valid_ids)
                    conn.execute(
                        f"""
                        UPDATE employees
                        SET advance_amount = (
                          SELECT COALESCE(SUM(amount), 0)
                          FROM employee_advances
                          WHERE company_id = ? AND employee_id = employees.id AND period = ?
                        )
                        WHERE company_id = ? AND id IN ({valid_placeholders})
                        """,
                        [company_id, period, company_id, *valid_ids],
                    )
                    audit_event(conn, company_id, user["email"], "bulk_update_employee_advance", "employees", None, None, {"ids": valid_ids, "advance_amount": advance_amount, "advance_date": advance_date})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Avans güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/site":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                employee_id = int(payload.get("employeeId", 0))
                site_name = normalize_text(payload.get("projectSiteName"))
                if employee_id <= 0:
                    self.send_json({"error": "Geçerli personel seçilmedi."}, 400)
                    return
                with connect() as conn:
                    before = conn.execute(
                        """
                        SELECT e.project_site_id, COALESCE(s.name, '') AS project_site
                        FROM employees e
                        LEFT JOIN project_sites s ON s.id = e.project_site_id
                        WHERE e.id = ? AND e.company_id = ?
                        """,
                        (employee_id, company_id),
                    ).fetchone()
                    if not before:
                        self.send_json({"error": "Personel bulunamadı."}, 404)
                        return
                    site_id = get_or_create_site(conn, site_name, company_id) if site_name else None
                    conn.execute(
                        "UPDATE employees SET project_site_id = ? WHERE id = ? AND company_id = ?",
                        (site_id, employee_id, company_id),
                    )
                    audit_event(conn, company_id, user["email"], "update_employee_site", "employees", employee_id, dict(before), {"project_site": site_name})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Şantiye güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/compensation":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                employee_id = int(payload.get("employeeId", 0))
                monthly_salary = to_float(payload.get("monthlySalary"))
                advance_amount = to_float(payload.get("advanceAmount"))
                advance_date = normalize_text(payload.get("advanceDate")) or date.today().isoformat()
                advance_note = normalize_text(payload.get("advanceNote")) or None
                iban = normalize_text(payload.get("iban")) or None
                leave_date = normalize_text(payload.get("leaveDate")) or None
                if employee_id <= 0:
                    self.send_json({"error": "Geçerli personel seçilmedi."}, 400)
                    return
                if monthly_salary < 0 or advance_amount < 0:
                    self.send_json({"error": "Maaş ve avans negatif olamaz."}, 400)
                    return
                with connect() as conn:
                    before = conn.execute(
                        "SELECT monthly_salary, advance_amount, iban_masked, leave_date FROM employees WHERE id = ? AND company_id = ?",
                        (employee_id, company_id),
                    ).fetchone()
                    if not before:
                        self.send_json({"error": "Personel bulunamadı."}, 404)
                        return
                    period = period_from_date(advance_date)
                    if advance_amount > 0:
                        conn.execute(
                            """
                            INSERT INTO employee_advances(company_id, employee_id, advance_date, period, amount, note)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """,
                            (company_id, employee_id, advance_date, period, advance_amount, advance_note),
                        )
                    conn.execute(
                        """
                        UPDATE employees
                        SET monthly_salary = ?,
                            advance_amount = (
                              SELECT COALESCE(SUM(amount), 0)
                              FROM employee_advances
                              WHERE company_id = ? AND employee_id = ? AND period = ?
                            ),
                            iban_masked = ?,
                            leave_date = ?
                        WHERE id = ? AND company_id = ?
                        """,
                        (monthly_salary, company_id, employee_id, period, iban, leave_date, employee_id, company_id),
                    )
                    audit_event(
                        conn,
                        company_id,
                        user["email"],
                        "update_compensation",
                        "employees",
                        employee_id,
                        dict(before),
                        {
                            "monthly_salary": monthly_salary,
                            "advance_amount": advance_amount,
                            "advance_date": advance_date if advance_amount > 0 else None,
                            "iban": iban,
                            "leave_date": leave_date,
                        },
                    )
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Personel güncellenemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/overtime":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                employee_id = int(payload.get("employeeId", 0))
                overtime_date = normalize_text(payload.get("overtimeDate")) or date.today().isoformat()
                hours = to_float(payload.get("hours"))
                hourly_rate = to_float(payload.get("hourlyRate"))
                amount = to_float(payload.get("amount")) or hours * hourly_rate
                note = normalize_text(payload.get("note")) or None
                if employee_id <= 0:
                    self.send_json({"error": "Geçerli personel seçilmedi."}, 400)
                    return
                if hours < 0 or hourly_rate < 0 or amount < 0:
                    self.send_json({"error": "Mesai saat, ücret ve tutar negatif olamaz."}, 400)
                    return
                if amount <= 0:
                    self.send_json({"error": "Mesai tutarı zorunlu."}, 400)
                    return
                with connect() as conn:
                    if not conn.execute("SELECT id FROM employees WHERE id = ? AND company_id = ?", (employee_id, company_id)).fetchone():
                        self.send_json({"error": "Personel bulunamadı."}, 404)
                        return
                    cur = conn.execute(
                        """
                        INSERT INTO employee_overtime_entries(company_id, employee_id, overtime_date, period, hours, hourly_rate, amount, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (company_id, employee_id, overtime_date, period_from_date(overtime_date), hours, hourly_rate, amount, note),
                    )
                    audit_event(conn, company_id, user["email"], "create_employee_overtime", "employee_overtime_entries", cur.lastrowid, None, payload)
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Mesai kaydedilemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/employees/delete":
            try:
                payload = read_json_body(self)
                company_id = self.selected_company_id(user, company_id_from_request(self, payload))
                ids = [int(item) for item in payload.get("ids", []) if int(item) > 0]
                if not ids:
                    employee_id = int(payload.get("employeeId", 0))
                    ids = [employee_id] if employee_id > 0 else []
                if not ids:
                    self.send_json({"error": "Silinecek personel seçilmedi."}, 400)
                    return
                placeholders = ",".join("?" for _ in ids)
                with connect() as conn:
                    conn.execute(
                        f"UPDATE employees SET status = 'deleted' WHERE company_id = ? AND id IN ({placeholders})",
                        [company_id, *ids],
                    )
                    audit_event(conn, company_id, user["email"], "delete_employees", "employees", None, None, {"ids": ids})
                self.send_json({"dashboard": dashboard_payload(user, company_id)})
            except json.JSONDecodeError:
                self.send_json({"error": "Geçersiz JSON."}, 400)
            except Exception as exc:
                self.send_json({"error": f"Personel silinemedi: {exc}"}, 500)
            return

        if parsed.path == "/api/attachments":
            ctype, _ = cgi.parse_header(self.headers.get("content-type"))
            if ctype != "multipart/form-data":
                self.send_json({"error": "PDF dosyası multipart/form-data olarak gönderilmeli."}, 400)
                return
            content_length = int(self.headers.get("content-length", "0") or 0)
            if content_length > MAX_UPLOAD_BYTES:
                self.send_json({"error": "Dosya boyutu izin verilen sınırı aşıyor."}, 413)
                return

            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("content-type"),
                },
            )
            company_id = self.selected_company_id(user, company_id_from_request(self, form=form))
            entity_type = normalize_text(form.getfirst("entityType"))
            entity_id = int(to_float(form.getfirst("entityId")))
            file_item = form["file"] if "file" in form else None
            if entity_type not in {"partner", "employee"} or entity_id <= 0:
                self.send_json({"error": "Geçerli kart seçilmedi."}, 400)
                return
            if file_item is None or not getattr(file_item, "filename", ""):
                self.send_json({"error": "PDF dosyası seçilmedi."}, 400)
                return

            original_name = Path(file_item.filename).name
            if not original_name.lower().endswith(".pdf"):
                self.send_json({"error": "Kart eklerine şimdilik yalnızca PDF yüklenebilir."}, 400)
                return

            table = "business_partners" if entity_type == "partner" else "employees"
            with connect() as conn:
                exists = conn.execute(f"SELECT id FROM {table} WHERE id = ? AND company_id = ?", (entity_id, company_id)).fetchone()
                if not exists:
                    self.send_json({"error": "Kart bulunamadı."}, 404)
                    return

            ATTACHMENTS.mkdir(parents=True, exist_ok=True)
            stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(8)}_{original_name}"
            saved = ATTACHMENTS / stored_name
            with saved.open("wb") as handle:
                shutil.copyfileobj(file_item.file, handle)

            with connect() as conn:
                cur = conn.execute(
                    """
                    INSERT INTO entity_attachments(company_id, entity_type, entity_id, file_name, stored_name, mime_type, file_size)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        company_id,
                        entity_type,
                        entity_id,
                        original_name,
                        stored_name,
                        getattr(file_item, "type", None) or "application/pdf",
                        saved.stat().st_size,
                    ),
                )
                audit_event(
                    conn,
                    company_id,
                    user["email"],
                    "upload_attachment",
                    "entity_attachments",
                    cur.lastrowid,
                    None,
                    {"entity_type": entity_type, "entity_id": entity_id, "file_name": original_name},
                )
            self.send_json({"dashboard": dashboard_payload(user, company_id)})
            return

        if parsed.path != "/api/import":
            self.send_json({"error": "Not found"}, 404)
            return

        ctype, pdict = cgi.parse_header(self.headers.get("content-type"))
        if ctype != "multipart/form-data":
            self.send_json({"error": "Excel dosyası multipart/form-data olarak gönderilmeli."}, 400)
            return
        content_length = int(self.headers.get("content-length", "0") or 0)
        if content_length > MAX_UPLOAD_BYTES:
            self.send_json({"error": "Dosya boyutu izin verilen sınırı aşıyor."}, 413)
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("content-type"),
            },
        )
        company_id = self.selected_company_id(user, company_id_from_request(self, form=form))
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            self.send_json({"error": "Dosya seçilmedi."}, 400)
            return

        original_name = Path(file_item.filename).name
        if not original_name.lower().endswith(".xlsx"):
            self.send_json({"error": "Şimdilik yalnızca .xlsx dosyası destekleniyor."}, 400)
            return

        UPLOADS.mkdir(parents=True, exist_ok=True)
        saved = UPLOADS / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        with saved.open("wb") as handle:
            shutil.copyfileobj(file_item.file, handle)

        try:
            summary = import_workbook(saved, original_name, company_id, user["email"])
        except Exception as exc:
            self.send_json({"error": f"Import sırasında hata oluştu: {exc}"}, 500)
            return

        self.send_json({"summary": summary, "dashboard": dashboard_payload(user, company_id)})


def main() -> int:
    init_db()
    port = int(os.environ.get("ERP_PORT", "8088"))
    server = ThreadingHTTPServer((HOST, port), ERPHandler)
    shown_host = "127.0.0.1" if HOST in ("0.0.0.0", "::") else HOST
    print(f"ERP taslak sunucusu hazır: http://{shown_host}:{port}")
    print("Kapatmak için Ctrl+C")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nSunucu kapatıldı.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
